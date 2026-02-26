#!/usr/bin/env python3
"""
build_tracker.py — Generates SPED_Time_Tracker.xlsx

Reads student names from 'Social Skills Times.xlsx' and creates a comprehensive
9-sheet SPED service tracking workbook with live Excel formulas.
"""

import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.pagebreak import Break
from openpyxl.workbook.defined_name import DefinedName

# ── Constants ────────────────────────────────────────────────────────────────

STUDENT_COUNT = 15
SESSIONS_PER_QUARTER = 10
QUARTERS = ["Q1", "Q2", "Q3", "Q4"]
QUARTER_LABELS = [
    "Q1 — Quarter 1", "Q2 — Quarter 2",
    "Q3 — Quarter 3", "Q4 — Quarter 4",
]

SERVICE_SHEETS = [
    ("Social Skills",     "4472C4"),   # blue
    ("Speech",            "6AA84F"),   # green
    ("OT",                "E69138"),   # orange
    ("Counseling",        "8E7CC3"),   # purple
    ("Inclusion Support", "C27BA0"),   # pink
]

# Colors (matching original workbook palette)
GRAY_TITLE   = "B7B7B7"
GREEN_HDR    = "57BB8A"
GREEN_GOOD   = "B7E1CD"
RED_BAD      = "EA9999"
YELLOW_WARN  = "FFD966"
NAVY         = "1F3864"
DARK_RED     = "8B0000"
YELLOW_TAB   = "FFD966"
WHITE        = "FFFFFF"
LIGHT_BLUE   = "D6E4F0"
HEADER_BG    = "D9E2F3"

# Reusable styles
TITLE_FONT     = Font(name="Arial", size=17, bold=True)
BOLD_WHITE_14  = Font(name="Arial", size=14, bold=True, color=WHITE)
BOLD_WHITE_13  = Font(name="Arial", size=13, bold=True, color=WHITE)
BOLD_WHITE_11  = Font(name="Arial", size=11, bold=True, color=WHITE)
COL_HDR_FONT   = Font(name="Arial", size=10, bold=True)
DATA_FONT      = Font(name="Arial", size=10)
BOLD_DATA_FONT = Font(name="Arial", size=10, bold=True)

TITLE_FILL   = PatternFill("solid", fgColor=GRAY_TITLE)
GREEN_FILL   = PatternFill("solid", fgColor=GREEN_GOOD)
RED_FILL     = PatternFill("solid", fgColor=RED_BAD)
YELLOW_FILL  = PatternFill("solid", fgColor=YELLOW_WARN)
NAVY_FILL    = PatternFill("solid", fgColor=NAVY)
HDR_BG_FILL  = PatternFill("solid", fgColor=HEADER_BG)
LBLUE_FILL   = PatternFill("solid", fgColor=LIGHT_BLUE)
DKRED_FILL   = PatternFill("solid", fgColor=DARK_RED)

THIN_BORDER = Border(
    left=Side("thin"), right=Side("thin"),
    top=Side("thin"),  bottom=Side("thin"),
)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _fill(color):
    return PatternFill("solid", fgColor=color)


def _set(cell, *, value=None, font=DATA_FONT, alignment=CENTER,
         border=THIN_BORDER, fill=None, fmt=None):
    """Helper to set multiple cell properties at once."""
    if value is not None:
        cell.value = value
    cell.font = font
    cell.alignment = alignment
    cell.border = border
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt


# ── Read source ──────────────────────────────────────────────────────────────

def read_students(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    names = []
    for r in range(3, 3 + STUDENT_COUNT):
        v = ws.cell(r, 1).value
        if v:
            names.append(str(v).strip())
    wb.close()
    return names


# ── Service sheets (5) ──────────────────────────────────────────────────────

def build_service_sheet(wb, name, color, students):
    ws = wb.create_sheet(title=name)
    ws.sheet_properties.tabColor = color

    # Title row
    ws.merge_cells("A1:AC1")
    _set(ws["A1"], value=f"{name} — SPED Minute Tracker",
         font=TITLE_FONT, fill=TITLE_FILL)

    # Column widths
    ws.column_dimensions["A"].width = 22
    for c in range(2, 22):
        ws.column_dimensions[get_column_letter(c)].width = 10
    for c in range(22, 30):
        ws.column_dimensions[get_column_letter(c)].width = 14

    # P / A / R dropdown
    pa_dv = DataValidation(
        type="list", formula1='"P,A,R"', allow_blank=True,
        showErrorMessage=True, errorTitle="Invalid",
        error="Enter P (Present), A (Absent), or R (Rescheduled)")
    ws.add_data_validation(pa_dv)

    # Minute / attendance column indices (1-based)
    min_cols = [2 + i * 2 for i in range(SESSIONS_PER_QUARTER)]      # B,D,F,...,T
    att_cols = [3 + i * 2 for i in range(SESSIONS_PER_QUARTER)]      # C,E,G,...,U
    min_ltrs = [get_column_letter(c) for c in min_cols]
    att_ltrs = [get_column_letter(c) for c in att_cols]

    # ── Build 4 quarterly sections ──
    for q in range(4):
        base   = 3 + q * 19           # quarter header row
        hdr    = base + 1             # column-header row
        first  = base + 2            # first student row
        last   = first + STUDENT_COUNT - 1

        # Quarter banner
        q_color = GREEN_HDR if q % 2 == 0 else color
        ws.merge_cells(f"A{base}:AC{base}")
        _set(ws.cell(base, 1), value=QUARTER_LABELS[q],
             font=BOLD_WHITE_14, fill=_fill(q_color))

        # Column headers
        headers = ["Student Name"]
        for s in range(1, SESSIONS_PER_QUARTER + 1):
            headers += [f"Sess {s}", f"Att {s}"]
        headers += ["Min Met", "Sess Att", "Absences", "IEP Req",
                     "Min Remain", "% Complete", "Pace Needed", "Projected"]

        for i, h in enumerate(headers, 1):
            _set(ws.cell(hdr, i), value=h, font=COL_HDR_FONT, fill=HDR_BG_FILL)

        # Student rows
        for si, sname in enumerate(students):
            r = first + si

            # A — name
            _set(ws.cell(r, 1), value=sname, font=DATA_FONT, alignment=LEFT)

            # Session + attendance cells
            for mc in min_cols:
                _set(ws.cell(r, mc), fmt="0")
            for ac in att_cols:
                cell = ws.cell(r, ac)
                _set(cell)
                pa_dv.add(cell)

            # V (22): Minutes Met
            refs = ",".join(f"{l}{r}" for l in min_ltrs)
            _set(ws.cell(r, 22), value=f"=SUM({refs})",
                 font=BOLD_DATA_FONT, fmt="0")

            # W (23): Sessions Attended  (count P's)
            cp = "+".join(f'COUNTIF({l}{r},"P")' for l in att_ltrs)
            _set(ws.cell(r, 23), value=f"={cp}", fmt="0")

            # X (24): Absences  (count A's)
            ca = "+".join(f'COUNTIF({l}{r},"A")' for l in att_ltrs)
            _set(ws.cell(r, 24), value=f"={ca}", fmt="0")

            # Y (25): IEP Required (user input)
            _set(ws.cell(r, 25), fmt="0")

            # Z (26): Minutes Remaining
            _set(ws.cell(r, 26),
                 value=f'=IF(Y{r}="","",Y{r}-V{r})', fmt="0")

            # AA (27): % Complete
            _set(ws.cell(r, 27),
                 value=f'=IF(OR(Y{r}="",Y{r}=0),"",V{r}/Y{r})',
                 font=BOLD_DATA_FONT, fmt="0%")

            # AB (28): Pace Needed
            _set(ws.cell(r, 28),
                 value=(f'=IF(OR(Y{r}="",Y{r}=0),"",IF((10-W{r}-X{r})<=0,'
                        f'"N/A",Z{r}/(10-W{r}-X{r})))'),
                 fmt="0.0")

            # AC (29): Projected Total
            _set(ws.cell(r, 29),
                 value=f'=IF(W{r}=0,"",ROUND((V{r}/W{r})*10,0))', fmt="0")

        # ── Conditional formatting per quarter ──
        rng = f"AA{first}:AA{last}"
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="greaterThanOrEqual", formula=["1"], fill=GREEN_FILL))
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="between", formula=["0.5", "0.9999"], fill=YELLOW_FILL))
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="lessThan", formula=["0.5"], fill=RED_FILL))

        v_rng = f"V{first}:V{last}"
        ws.conditional_formatting.add(v_rng, FormulaRule(
            formula=[f"AND(V{first}>=Y{first},Y{first}<>\"\")"], fill=GREEN_FILL))
        ws.conditional_formatting.add(v_rng, FormulaRule(
            formula=[f'AND(V{first}<Y{first},Y{first}<>"")'], fill=RED_FILL))

        ab_rng = f"AB{first}:AB{last}"
        ws.conditional_formatting.add(ab_rng, FormulaRule(
            formula=[f"AND(ISNUMBER(AB{first}),W{first}>0,"
                     f"AB{first}>2*(V{first}/W{first}))"],
            fill=RED_FILL))

        # Consecutive-absence highlight on student name
        checks = [f'AND({att_ltrs[i]}{first}="A",{att_ltrs[i+1]}{first}="A")'
                  for i in range(len(att_ltrs) - 1)]
        ws.conditional_formatting.add(
            f"A{first}:A{last}",
            FormulaRule(formula=[f"OR({','.join(checks)})"], fill=RED_FILL))

    # Freeze panes & print
    ws.freeze_panes = "B5"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


# ── Session Notes ────────────────────────────────────────────────────────────

def build_session_notes(wb, students):
    ws = wb.create_sheet(title="Session Notes")
    ws.sheet_properties.tabColor = YELLOW_TAB

    ws.merge_cells("A1:G1")
    _set(ws["A1"], value="Session Notes", font=TITLE_FONT, fill=TITLE_FILL)

    headers = ["Date", "Student Name", "Service Type", "Skill Targeted",
               "Student Response", "Behavior Notes", "Provider"]
    widths  = [14, 22, 18, 22, 30, 30, 18]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        _set(ws.cell(2, i), value=h, font=BOLD_WHITE_11,
             fill=_fill("4472C4"))
        ws.column_dimensions[get_column_letter(i)].width = w

    # Data validations
    name_list = ",".join(students)
    name_dv = DataValidation(type="list", formula1=f'"{name_list}"',
                             allow_blank=True)
    svc_list = ",".join(s[0] for s in SERVICE_SHEETS)
    svc_dv  = DataValidation(type="list", formula1=f'"{svc_list}"',
                             allow_blank=True)
    ws.add_data_validation(name_dv)
    ws.add_data_validation(svc_dv)

    for r in range(3, 503):
        for c in range(1, 8):
            cell = ws.cell(r, c)
            _set(cell, alignment=LEFT if c >= 4 else CENTER,
                 fill=LBLUE_FILL if r % 2 == 1 else None)
        ws.cell(r, 1).number_format = "MM/DD/YYYY"
        name_dv.add(ws.cell(r, 2))
        svc_dv.add(ws.cell(r, 3))

    ws.freeze_panes = "A3"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


# ── Year Summary ─────────────────────────────────────────────────────────────

def build_year_summary(wb, students):
    ws = wb.create_sheet(title="Year Summary")
    ws.sheet_properties.tabColor = NAVY

    ws.merge_cells("A1:N1")
    _set(ws["A1"], value="Year Summary — SPED Minute Tracker",
         font=Font(name="Arial", size=17, bold=True, color=WHITE), fill=NAVY_FILL)

    ws.column_dimensions["A"].width = 22
    for c in range(2, 15):
        ws.column_dimensions[get_column_letter(c)].width = 14

    block_row = 3
    for svc_idx, (svc_name, svc_color) in enumerate(SERVICE_SHEETS):
        # Service header
        ws.merge_cells(f"A{block_row}:N{block_row}")
        _set(ws.cell(block_row, 1), value=svc_name,
             font=BOLD_WHITE_13, fill=_fill(svc_color))

        col_row = block_row + 1
        col_headers = [
            "Student Name",
            "Q1 Met", "Q2 Met", "Q3 Met", "Q4 Met", "Year Total",
            "Q1 IEP", "Q2 IEP", "Q3 IEP", "Q4 IEP", "Annual IEP",
            "Year %", "Total Sess", "Total Abs",
        ]
        for i, h in enumerate(col_headers, 1):
            _set(ws.cell(col_row, i), value=h, font=COL_HDR_FONT, fill=HDR_BG_FILL)

        svc = f"'{svc_name}'"
        for si in range(STUDENT_COUNT):
            r = col_row + 1 + si
            _set(ws.cell(r, 1), value=students[si], alignment=LEFT)

            for q in range(4):
                svc_r = 5 + q * 19 + si   # data row on service sheet

                # Qn Met  (cols 2-5) ← service col V (22)
                _set(ws.cell(r, 2 + q), value=f"={svc}!V{svc_r}", fmt="0")
                # Qn IEP  (cols 7-10) ← service col Y (25)
                _set(ws.cell(r, 7 + q), value=f"={svc}!Y{svc_r}", fmt="0")

            # F (6): Year Total
            _set(ws.cell(r, 6), value=f"=SUM(B{r}:E{r})",
                 font=BOLD_DATA_FONT, fmt="0")

            # K (11): Annual IEP
            _set(ws.cell(r, 11), value=f"=SUM(G{r}:J{r})",
                 font=BOLD_DATA_FONT, fmt="0")

            # L (12): Year %
            _set(ws.cell(r, 12),
                 value=f'=IF(K{r}=0,"",F{r}/K{r})',
                 font=BOLD_DATA_FONT, fmt="0%")

            # M (13): Total Sessions
            sess = "+".join(f"{svc}!W{5 + q * 19 + si}" for q in range(4))
            _set(ws.cell(r, 13), value=f"={sess}", fmt="0")

            # N (14): Total Absences
            abss = "+".join(f"{svc}!X{5 + q * 19 + si}" for q in range(4))
            _set(ws.cell(r, 14), value=f"={abss}", fmt="0")

        # Conditional formatting on Year %
        pct_rng = f"L{col_row + 1}:L{col_row + STUDENT_COUNT}"
        ws.conditional_formatting.add(pct_rng, CellIsRule(
            operator="greaterThanOrEqual", formula=["0.8"], fill=GREEN_FILL))
        ws.conditional_formatting.add(pct_rng, CellIsRule(
            operator="between", formula=["0.5", "0.7999"], fill=YELLOW_FILL))
        ws.conditional_formatting.add(pct_rng, CellIsRule(
            operator="lessThan", formula=["0.5"], fill=RED_FILL))

        block_row = col_row + STUDENT_COUNT + 2

    ws.freeze_panes = "B3"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


# ── Dashboard ────────────────────────────────────────────────────────────────

def build_dashboard(wb, students):
    ws = wb.create_sheet(title="Dashboard", index=0)
    ws.sheet_properties.tabColor = NAVY

    ws.merge_cells("A1:G1")
    _set(ws["A1"], value="SPED Service Compliance Dashboard",
         font=Font(name="Arial", size=18, bold=True, color=WHITE), fill=NAVY_FILL)

    ws.merge_cells("A2:G2")
    _set(ws["A2"], value="Auto-calculated from Year Summary data",
         font=Font(name="Arial", size=11, italic=True, color="666666"),
         border=None)

    dash_hdrs = ["Service Type", "Total Students", "On Track (≥80%)",
                 "Behind (50-79%)", "At Risk (<50%)", "Overall %", "Status"]
    widths    = [22, 16, 18, 18, 18, 14, 14]
    for i, (h, w) in enumerate(zip(dash_hdrs, widths), 1):
        _set(ws.cell(4, i), value=h, font=BOLD_WHITE_11, fill=NAVY_FILL)
        ws.column_dimensions[get_column_letter(i)].width = w

    ys = "'Year Summary'"
    for si, (svc_name, svc_color) in enumerate(SERVICE_SHEETS):
        r = 5 + si
        ys_s = 5 + si * 18                     # first data row in Year Summary
        ys_e = ys_s + STUDENT_COUNT - 1         # last data row

        _set(ws.cell(r, 1), value=svc_name,
             font=Font(name="Arial", size=11, bold=True, color=WHITE),
             fill=_fill(svc_color), alignment=LEFT)

        # B: Total Students with data
        _set(ws.cell(r, 2),
             value=f"=COUNTA({ys}!L{ys_s}:L{ys_e})",
             font=Font(name="Arial", size=12, bold=True))

        # C: On Track (>=80%)
        _set(ws.cell(r, 3),
             value=f'=COUNTIF({ys}!L{ys_s}:L{ys_e},">="&0.8)')

        # D: Behind (50-79%)
        _set(ws.cell(r, 4),
             value=(f'=COUNTIFS({ys}!L{ys_s}:L{ys_e},">="&0.5,'
                    f'{ys}!L{ys_s}:L{ys_e},"<"&0.8)'))

        # E: At Risk (<50%)
        _set(ws.cell(r, 5),
             value=f'=COUNTIF({ys}!L{ys_s}:L{ys_e},"<"&0.5)')

        # F: Overall %
        _set(ws.cell(r, 6),
             value=(f'=IF(COUNTA({ys}!L{ys_s}:L{ys_e})=0,"",'
                    f'AVERAGE({ys}!L{ys_s}:L{ys_e}))'),
             font=Font(name="Arial", size=11, bold=True), fmt="0%")

        # G: Status
        _set(ws.cell(r, 7),
             value=(f'=IF(F{r}="","",IF(F{r}>=0.8,"On Track",'
                    f'IF(F{r}>=0.5,"Behind","At Risk")))'),
             font=BOLD_DATA_FONT)

        # Conditional formatting per row
        ws.conditional_formatting.add(f"C{r}", CellIsRule(
            operator="greaterThan", formula=["0"], fill=GREEN_FILL))
        ws.conditional_formatting.add(f"D{r}", CellIsRule(
            operator="greaterThan", formula=["0"], fill=YELLOW_FILL))
        ws.conditional_formatting.add(f"E{r}", CellIsRule(
            operator="greaterThan", formula=["0"], fill=RED_FILL))

        for col_ltr, thresholds in [("F", [(0.8, GREEN_FILL),
                                            (0.5, YELLOW_FILL)])]:
            ws.conditional_formatting.add(f"{col_ltr}{r}", CellIsRule(
                operator="greaterThanOrEqual",
                formula=[str(thresholds[0][0])], fill=thresholds[0][1]))
            ws.conditional_formatting.add(f"{col_ltr}{r}", CellIsRule(
                operator="between",
                formula=[str(thresholds[1][0]), "0.7999"],
                fill=thresholds[1][1]))
            ws.conditional_formatting.add(f"{col_ltr}{r}", CellIsRule(
                operator="lessThan",
                formula=[str(thresholds[1][0])], fill=RED_FILL))

        ws.conditional_formatting.add(f"G{r}", FormulaRule(
            formula=[f'G{r}="On Track"'], fill=GREEN_FILL))
        ws.conditional_formatting.add(f"G{r}", FormulaRule(
            formula=[f'G{r}="Behind"'], fill=YELLOW_FILL))
        ws.conditional_formatting.add(f"G{r}", FormulaRule(
            formula=[f'G{r}="At Risk"'], fill=RED_FILL))

    # Summary row
    sr = 5 + len(SERVICE_SHEETS) + 1
    _set(ws.cell(sr, 1), value="ALL SERVICES",
         font=Font(name="Arial", size=11, bold=True, color=WHITE),
         fill=_fill("333333"), alignment=LEFT)
    for c in range(2, 6):
        ltr = get_column_letter(c)
        _set(ws.cell(sr, c),
             value=f"=SUM({ltr}5:{ltr}{sr - 2})",
             font=Font(name="Arial", size=11, bold=True))
    _set(ws.cell(sr, 6),
         value=f'=IF(COUNT(F5:F{sr - 2})=0,"",AVERAGE(F5:F{sr - 2}))',
         font=Font(name="Arial", size=11, bold=True), fmt="0%")

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


# ── ARD Report ───────────────────────────────────────────────────────────────

def build_ard_report(wb, students):
    ws = wb.create_sheet(title="ARD Report")
    ws.sheet_properties.tabColor = DARK_RED

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14

    cur = 1
    ys = "'Year Summary'"
    ard_hdrs = ["Service", "IEP Required", "Minutes Delivered",
                "% Complete", "Sessions Attended", "Absences", "Status"]

    for si, sname in enumerate(students):
        # Student header
        ws.merge_cells(f"A{cur}:G{cur}")
        _set(ws.cell(cur, 1), value=sname,
             font=Font(name="Arial", size=14, bold=True, color=WHITE),
             fill=DKRED_FILL)
        cur += 1

        # Column headers
        for i, h in enumerate(ard_hdrs, 1):
            _set(ws.cell(cur, i), value=h, font=COL_HDR_FONT,
                 fill=_fill("D9D9D9"))
        cur += 1

        svc_start = cur
        for svc_idx, (svc_name, _) in enumerate(SERVICE_SHEETS):
            r = cur
            ys_r = 5 + svc_idx * 18 + si   # Year Summary data row

            _set(ws.cell(r, 1), value=svc_name, font=BOLD_DATA_FONT,
                 alignment=LEFT)
            _set(ws.cell(r, 2), value=f"={ys}!K{ys_r}", fmt="0")        # IEP Req
            _set(ws.cell(r, 3), value=f"={ys}!F{ys_r}", fmt="0")        # Delivered
            _set(ws.cell(r, 4), value=f"={ys}!L{ys_r}",                 # % Compl
                 font=BOLD_DATA_FONT, fmt="0%")
            _set(ws.cell(r, 5), value=f"={ys}!M{ys_r}", fmt="0")        # Sess Att
            _set(ws.cell(r, 6), value=f"={ys}!N{ys_r}", fmt="0")        # Absences
            _set(ws.cell(r, 7),                                          # Status
                 value=(f'=IF(D{r}="","",IF(D{r}>=0.8,"On Track",'
                        f'IF(D{r}>=0.5,"Behind","At Risk")))'),
                 font=BOLD_DATA_FONT)
            cur += 1

        svc_end = cur - 1

        # Conditional formatting — % Complete
        ws.conditional_formatting.add(f"D{svc_start}:D{svc_end}", CellIsRule(
            operator="greaterThanOrEqual", formula=["0.8"], fill=GREEN_FILL))
        ws.conditional_formatting.add(f"D{svc_start}:D{svc_end}", CellIsRule(
            operator="between", formula=["0.5", "0.7999"], fill=YELLOW_FILL))
        ws.conditional_formatting.add(f"D{svc_start}:D{svc_end}", CellIsRule(
            operator="lessThan", formula=["0.5"], fill=RED_FILL))

        # Conditional formatting — Status
        ws.conditional_formatting.add(f"G{svc_start}:G{svc_end}", FormulaRule(
            formula=[f'G{svc_start}="On Track"'], fill=GREEN_FILL))
        ws.conditional_formatting.add(f"G{svc_start}:G{svc_end}", FormulaRule(
            formula=[f'G{svc_start}="Behind"'], fill=YELLOW_FILL))
        ws.conditional_formatting.add(f"G{svc_start}:G{svc_end}", FormulaRule(
            formula=[f'G{svc_start}="At Risk"'], fill=RED_FILL))

        # Page break between students
        if si < STUDENT_COUNT - 1:
            cur += 1   # blank separator row
            ws.row_breaks.append(Break(id=cur - 1))

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    src = "Social Skills Times.xlsx"
    out = "SPED_Time_Tracker.xlsx"

    if not os.path.exists(src):
        print(f"Error: '{src}' not found in current directory.")
        return

    students = read_students(src)
    print(f"Read {len(students)} students: {', '.join(students)}")

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    print("Building Dashboard...")
    build_dashboard(wb, students)

    print("Building service sheets...")
    for svc_name, svc_color in SERVICE_SHEETS:
        print(f"  - {svc_name}")
        build_service_sheet(wb, svc_name, svc_color, students)

    print("Building Session Notes...")
    build_session_notes(wb, students)

    print("Building Year Summary...")
    build_year_summary(wb, students)

    print("Building ARD Report...")
    build_ard_report(wb, students)

    # Named range for student names (references Social Skills Q1 names)
    dn = DefinedName("StudentNames",
                     attr_text="'Social Skills'!$A$5:$A$19")
    wb.defined_names.add(dn)

    wb.save(out)
    print(f"\nCreated: {out}")


if __name__ == "__main__":
    main()
